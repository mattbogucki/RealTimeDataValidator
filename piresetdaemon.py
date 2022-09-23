import argparse
import datetime
import os
import socket
import sys

import PIconnect
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

BIG_NUMBER = 99999999999999


def main():
    parser = argparse.ArgumentParser(description='Points List Validator')
    parser.add_argument('FILENAME', action="store", help=".xlsx filename of points list")
    parser.add_argument('TAB', action="store", help='Name of tab with owners points list (Default is DNP3.0 Points List)')
    parser.add_argument('-d', action="store_true", dest="debug", default=False)

    try:
        cmd_line_args = parser.parse_args()
    except:
        print('Supply the proper arguments')
        sys.exit(-1)

    try:
        test_wb = openpyxl.load_workbook(cmd_line_args.FILENAME, data_only=True)
        test_sheet = test_wb[cmd_line_args.TAB]
    except:
        print('File Name and Sheet not valid')
        sys.exit(-2)

    desktop_logfile = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop\point_rt_validation_log.txt')
    pm = PointsMaster(cmd_line_args.FILENAME, cmd_line_args.TAB, desktop_logfile)
    pm.validate_rt_pi_data()


def get_pi_server_name() -> str:
    try:
        pi_servers = list(PIconnect.PIServer.servers.keys())
        for pi_server in pi_servers:
            if "azw2" in pi_server.lower():
                continue
            return pi_server
        return socket.gethostname()
    except:
        return socket.gethostname()


# TODO Could make this even more fitting to the data size, but good enough for now
def size_columns_to_be_readable(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        column_letter = get_column_letter(col)
        column_width = 70 if column_letter == "A" else 25
        dim_holder[column_letter] = ColumnDimension(ws, min=col, max=col, width=column_width)

    ws.column_dimensions = dim_holder


class PointsMaster(object):

    def __init__(self, workbook, sheet,
                 OUTPUT_FILE="errors.txt",
                 DEVICE_ID=1,
                 POINT_NAME=2,
                 DEVICE_TYPE=3,
                 DESCRIPTION=5,
                 AVAILABLE=8,
                 start=20):
        self.workbook = workbook
        self.sheet = sheet
        self.start = start
        self.HEADER = 0
        self.DNP_INDEX = 6
        self.STATE_TABLE = 7
        self.UNITS = 7
        self.SUBSTATION = 0
        self.EGU_MIN = 9
        self.EGU_MAX = 10
        self.POINT_NAME = POINT_NAME
        self.DEVICE_TYPE = DEVICE_TYPE
        self.DEVICE_ID = DEVICE_ID
        self.AVAILABLE = AVAILABLE
        self.DESCRIPTION = DESCRIPTION
        self.OUTPUT_FILE = OUTPUT_FILE

    def generate_dictionary_of_points_from_attachment3(self) -> dict:
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        is_analog_point = False
        point_metadatata_dict = {}

        for i, row in enumerate(wb[self.sheet]):
            point_type_header = str(row[self.HEADER].value).strip()
            if point_type_header == "Digital Inputs" or point_type_header == "Digital Outputs" or \
                    point_type_header == "Counters" or point_type_header == "Analog Outputs":
                is_analog_point = False
            elif point_type_header == "Analog Inputs":
                is_analog_point = True
            if is_analog_point:
                availability = str(row[self.AVAILABLE].value).strip()
                if availability == "Requested-Available" or availability == 'Not Requested-Available':
                    point_name = str(row[self.POINT_NAME].value).strip()
                    device_type = str(row[self.DEVICE_TYPE].value).strip()
                    source_device = str(row[self.DEVICE_ID].value).strip()
                    egu_min_string = row[self.EGU_MIN].value
                    egu_max_string = row[self.EGU_MAX].value
                    try:
                        egu_min = float(egu_min_string)
                        egu_max = float(egu_max_string)
                    except:
                        #print("ERROR: Row {} - Invalid Engineering Units Min:{} Max:{}, Must be numeric".format(i + 1, egu_min_string, egu_max_string))
                        continue

                    point_metadatata_dict[point_name] = (device_type, source_device, egu_max, egu_min)
        print("A11 Point Count: {}".format(len(point_metadatata_dict)))
        return point_metadatata_dict

    def validate_rt_pi_data(self):
        point_metadata_dict = self.generate_dictionary_of_points_from_attachment3()
        pi_server = get_pi_server_name()
        print(pi_server)
        points_audited = 0
        min_violations = []
        max_violations = []
        frequency_of_update_violations = []
        granularity_violations = []

        with PIconnect.PIServer(server=pi_server) as server:
            all_points = server.search('*')
            print("Total Points on Pi Server:{}".format(len(all_points)))
            for point, values in point_metadata_dict.items():
                device_type = values[0]
                source_device = values[1]
                egu_max = values[2]
                egu_min = values[3]
                try:
                    this_point = server.search(str(point))[0]
                    data = this_point.recorded_values('*-24h', '*')
                except Exception as e:
                    #print("ERROR: Could not find {} on the PI server".format(point))
                    continue

                points_audited += 1
                recorded_values_count = len(data)
                max_recorded_value = max(data)
                min_recorded_value = min(data)

                print("{} Count:{} Min:{} Max:{}".format(point, recorded_values_count, min_recorded_value, max_recorded_value))

                # Check to ensure all data falls within min/max from attachment 3
                if max_recorded_value > egu_max:
                    max_violations.append((point, egu_max, max_recorded_value, device_type, source_device))
                if min_recorded_value < egu_min:
                    min_violations.append((point, egu_min, min_recorded_value, device_type, source_device))

                # Should be a value at least every 5 minutes at a bare minimum
                if recorded_values_count < 288:
                    frequency_of_update_violations.append((point, recorded_values_count, device_type, source_device))

                smallest_delta = BIG_NUMBER

                # Should record deltas at a minimum of 0.1
                for i in range(recorded_values_count):
                    if i == recorded_values_count-1:
                        break
                    previous_value = data[i]
                    current_value = data[i+1]

                    delta = abs(current_value - previous_value)
                    if delta < smallest_delta and delta != 0:  # 0 doesnt count
                        smallest_delta = delta

                if smallest_delta >= 1:
                    granularity_violations.append((point, smallest_delta, device_type, source_device))

        print("Total Points Audited: {}".format(points_audited))

        # Sort by worst offenders
        min_violations.sort(key=lambda x: (x[1] - x[2]), reverse=True)
        max_violations.sort(key=lambda x: (x[2] - x[1]), reverse=True)
        frequency_of_update_violations.sort(key=lambda x: x[1], reverse=False)
        granularity_violations.sort(key=lambda x: x[1], reverse=True)

        excel_file_output = openpyxl.Workbook()
        excel_file_output.title = "Real-Time-Data-Audit"
        excel_file_output.active.title = "Min Violations"
        excel_file_output.create_sheet("Max Violations")
        excel_file_output.create_sheet("Update Freq Violations")
        excel_file_output.create_sheet("Granularity Violations")

        min_violations_ws = excel_file_output["Min Violations"]
        min_violations_ws.cell(row=1, column=1).value = "Point Name"
        min_violations_ws.cell(row=1, column=2).value = "A11 EGU Min"
        min_violations_ws.cell(row=1, column=3).value = "Recorded Min"
        min_violations_ws.cell(row=1, column=4).value = "Device Type"
        min_violations_ws.cell(row=1, column=5).value = "Source Device"
        for i, min_violation in enumerate(min_violations, start=2):  # Start after header row openpyxl is 1 based
            min_violations_ws.cell(row=i, column=1).value = min_violation[0]
            min_violations_ws.cell(row=i, column=2).value = min_violation[1]
            min_violations_ws.cell(row=i, column=3).value = min_violation[2]
            min_violations_ws.cell(row=i, column=4).value = min_violation[3]
            min_violations_ws.cell(row=i, column=5).value = min_violation[4]

        max_violations_ws = excel_file_output["Max Violations"]
        max_violations_ws.cell(row=1, column=1).value = "Point Name"
        max_violations_ws.cell(row=1, column=2).value = "A11 EGU Max"
        max_violations_ws.cell(row=1, column=3).value = "Recorded Max"
        max_violations_ws.cell(row=1, column=4).value = "Device Type"
        max_violations_ws.cell(row=1, column=5).value = "Source Device"
        for j, max_violation in enumerate(max_violations, start=2):
            max_violations_ws.cell(row=j, column=1).value = max_violation[0]
            max_violations_ws.cell(row=j, column=2).value = max_violation[1]
            max_violations_ws.cell(row=j, column=3).value = max_violation[2]
            max_violations_ws.cell(row=j, column=4).value = max_violation[3]
            max_violations_ws.cell(row=j, column=5).value = max_violation[4]

        freq_violations_ws = excel_file_output["Update Freq Violations"]
        freq_violations_ws.cell(row=1, column=1).value = "Point Name"
        freq_violations_ws.cell(row=1, column=2).value = "Updates in last 24 hours"
        freq_violations_ws.cell(row=1, column=3).value = "Device Type"
        freq_violations_ws.cell(row=1, column=4).value = "Source Device"
        for k, frequency_of_update_violation in enumerate(frequency_of_update_violations, start=2):
            freq_violations_ws.cell(row=k, column=1).value = frequency_of_update_violation[0]
            freq_violations_ws.cell(row=k, column=2).value = frequency_of_update_violation[1]
            freq_violations_ws.cell(row=k, column=3).value = frequency_of_update_violation[2]
            freq_violations_ws.cell(row=k, column=4).value = frequency_of_update_violation[3]

        granularity_violations_ws = excel_file_output["Granularity Violations"]
        granularity_violations_ws.cell(row=1, column=1).value = "Point Name"
        granularity_violations_ws.cell(row=1, column=2).value = "Smallest Granularity Change in Last 24 hours"
        granularity_violations_ws.cell(row=1, column=3).value = "Device Type"
        granularity_violations_ws.cell(row=1, column=4).value = "Source Device"
        for m, granularity_violation in enumerate(granularity_violations, start=2):
            granularity_violations_ws.cell(row=m, column=1).value = granularity_violation[0]
            if granularity_violation[1] == BIG_NUMBER:
                granularity_violations_ws.cell(row=m, column=2).value = "Not Updating"
            else:
                granularity_violations_ws.cell(row=m, column=2).value = granularity_violation[1]
            granularity_violations_ws.cell(row=m, column=3).value = granularity_violation[2]
            granularity_violations_ws.cell(row=m, column=4).value = granularity_violation[3]

        for sheet in excel_file_output.sheetnames:
            ws = excel_file_output[sheet]
            size_columns_to_be_readable(ws)

        now = datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        excel_file_output.save("{}-Real-Time-Data-Audit_{}.xlsx".format(get_pi_server_name(), now))


if __name__ == "__main__":
    main()
